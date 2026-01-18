#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script to collect French small business data with owner information.
Searches Google Maps for businesses in France and exports to Excel.

Usage:
    python scripts/collect_france_businesses.py --category "salon de beauté" --location "Paris"
    python scripts/collect_france_businesses.py --categories "coiffeur" "restaurant" --location "Lyon" --max-results 50
"""

import asyncio
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

import httpx
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# Add parent directory to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from config import settings
from utils.lead_generation.google_maps import GoogleMapsClient
from utils.logging_config import setup_logging
from scripts.browser_scraper import BrowserScraper

logger = setup_logging(
    name=__name__, log_level="INFO", log_file="france_businesses.log", log_dir="logs"
)


class FranceBusinessCollector:
    """Collector for French business data with owner information."""

    def __init__(
        self,
        google_maps_api_key: Optional[str] = None,
        output_dir: str = "research_results/france",
        use_browser_fallback: bool = True,
    ):
        """
        Initialize collector.

        Args:
            google_maps_api_key: Google Maps API key (optional, uses env var if not provided)
            output_dir: Directory to save results
            use_browser_fallback: Whether to use browser scraper as fallback
        """
        browser_scraper = BrowserScraper(use_mcp_puppeteer=True) if use_browser_fallback else None
        self.google_maps = GoogleMapsClient(
            api_key=google_maps_api_key,
            browser_scraper=browser_scraper,
        )
        self.browser_scraper = browser_scraper
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.http_client = httpx.AsyncClient(timeout=30.0, follow_redirects=True)

    async def close(self):
        """Close all clients."""
        await self.google_maps.close()
        await self.http_client.aclose()
        if self.browser_scraper:
            await self.browser_scraper.close()

    async def search_businesses(
        self,
        categories: List[str],
        location: str = "France",
        max_per_category: int = 20,
    ) -> List[Dict]:
        """
        Search for businesses using Google Maps.

        Args:
            categories: List of business categories to search
            location: Location to search in (default: France)
            max_per_category: Maximum results per category

        Returns:
            List of business data dictionaries
        """
        all_businesses = []

        for category in categories:
            logger.info(f"Searching for '{category}' in {location}...")

            try:
                # Search Google Maps
                places = await self.google_maps.search_businesses(
                    query=category, location=location, max_results=max_per_category
                )

                logger.info(f"Found {len(places)} places for '{category}'")

                # Process each place
                for place_data in places:
                    try:
                        business = await self._enrich_business(place_data, category)
                        if business:
                            all_businesses.append(business)
                            logger.info(
                                f"✓ Processed: {business.get('business_name', 'Unknown')} "
                                f"(Phone: {business.get('phone', 'N/A')}, "
                                f"Owner: {business.get('owner_name', 'N/A')})"
                            )

                        # Rate limiting
                        await asyncio.sleep(1)

                    except Exception as e:
                        logger.error(f"Error processing place: {e}", exc_info=True)
                        continue

            except Exception as e:
                logger.error(f"Error searching category '{category}': {e}", exc_info=True)
                continue

        return all_businesses

    async def _enrich_business(
        self, place_data: Dict, category: str
    ) -> Optional[Dict]:
        """
        Enrich Google Maps place data with owner information.

        Args:
            place_data: Raw Google Maps place data
            category: Business category

        Returns:
            Enriched business dictionary or None
        """
        # Parse Google Maps data
        google_info = self.google_maps.parse_place_data(place_data)

        business = {
            "business_name": google_info.get("business_name", ""),
            "address": google_info.get("google_address", ""),
            "phone": google_info.get("google_phone", ""),
            "email": None,
            "website": google_info.get("google_website", ""),
            "category": category,
            "google_maps_url": google_info.get("google_maps_url", ""),
            "owner_name": None,
            "owner_first_name": None,
            "owner_last_name": None,
        }

        # Try to find owner information
        owner_info = await self._find_owner_info(
            business["business_name"], business["address"], business["website"]
        )
        if owner_info:
            business.update(owner_info)

        # Try to find email from website
        if business["website"] and not business["email"]:
            email = await self._extract_email_from_website(business["website"])
            if email:
                business["email"] = email

        return business

    async def _find_owner_info(
        self, business_name: str, address: str, website: Optional[str]
    ) -> Optional[Dict]:
        """
        Find business owner information using web search.

        Args:
            business_name: Business name
            address: Business address
            website: Business website URL

        Returns:
            Dictionary with owner information or None
        """
        # Try multiple search strategies
        search_queries = [
            f"{business_name} propriétaire",
            f"{business_name} directeur",
            f"{business_name} gérant",
            f"{business_name} owner",
            f"{business_name} Infogreffe",
        ]

        for query in search_queries:
            try:
                # Use DuckDuckGo or Google search
                owner_info = await self._search_owner_web(query, business_name)
                if owner_info:
                    return owner_info
            except Exception as e:
                logger.debug(f"Owner search failed for query '{query}': {e}")
                continue

        # Try Infogreffe search if we have business name
        if business_name:
            try:
                owner_info = await self._search_infogreffe(business_name)
                if owner_info:
                    return owner_info
            except Exception as e:
                logger.debug(f"Infogreffe search failed: {e}")

        return None

    async def _search_owner_web(
        self, query: str, business_name: str
    ) -> Optional[Dict]:
        """
        Search for owner information using web search.

        Args:
            query: Search query
            business_name: Business name for context

        Returns:
            Dictionary with owner information or None
        """
        # Use DuckDuckGo HTML search (no API key needed)
        search_url = f"https://html.duckduckgo.com/html/?q={query.replace(' ', '+')}"

        try:
            response = await self.http_client.get(search_url)
            response.raise_for_status()

            soup = BeautifulSoup(response.text, "html.parser")

            # Look for owner information in search results
            # This is a simplified extraction - could be improved
            results = soup.find_all("a", class_="result__a")
            for result in results[:5]:  # Check first 5 results
                text = result.get_text()
                # Look for common French owner patterns
                patterns = [
                    r"gérant[:\s]+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)",
                    r"directeur[:\s]+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)",
                    r"propriétaire[:\s]+([A-Z][a-z]+(?:\s+[A-Z][a-z]+)*)",
                ]

                for pattern in patterns:
                    match = re.search(pattern, text, re.IGNORECASE)
                    if match:
                        owner_name = match.group(1).strip()
                        # Split into first and last name
                        name_parts = owner_name.split()
                        return {
                            "owner_name": owner_name,
                            "owner_first_name": name_parts[0] if name_parts else None,
                            "owner_last_name": " ".join(name_parts[1:]) if len(name_parts) > 1 else None,
                        }

        except Exception as e:
            logger.debug(f"Web search failed: {e}")

        return None

    async def _search_infogreffe(self, business_name: str) -> Optional[Dict]:
        """
        Search Infogreffe (French business registry) for owner information.

        Args:
            business_name: Business name to search

        Returns:
            Dictionary with owner information or None
        """
        # Infogreffe search URL
        search_url = "https://www.infogreffe.fr/infogreffe/consulter-entreprise/recherche-entreprise.html"

        try:
            # Note: Infogreffe requires form submission, this is a simplified approach
            # For production, would need to handle form submission and CAPTCHA
            params = {"denomination": business_name}
            response = await self.http_client.get(search_url, params=params)
            response.raise_for_status()

            soup = BeautifulSoup(response.text, "html.parser")

            # Extract owner information from Infogreffe page
            # This is simplified - actual implementation would need to parse the results page
            # Look for "Gérant" or "Dirigeant" sections
            gérant_section = soup.find(string=re.compile(r"Gérant|Dirigeant", re.I))
            if gérant_section:
                # Try to extract name from nearby text
                parent = gérant_section.find_parent()
                if parent:
                    text = parent.get_text()
                    name_match = re.search(
                        r"([A-Z][a-z]+(?:\s+[A-Z][a-z]+)+)", text
                    )
                    if name_match:
                        owner_name = name_match.group(1).strip()
                        name_parts = owner_name.split()
                        return {
                            "owner_name": owner_name,
                            "owner_first_name": name_parts[0] if name_parts else None,
                            "owner_last_name": " ".join(name_parts[1:]) if len(name_parts) > 1 else None,
                        }

        except Exception as e:
            logger.debug(f"Infogreffe search failed: {e}")

        return None

    async def _extract_email_from_website(self, website: str) -> Optional[str]:
        """
        Extract email address from business website.

        Args:
            website: Website URL

        Returns:
            Email address or None
        """
        try:
            response = await self.http_client.get(website)
            response.raise_for_status()

            # Look for email patterns in HTML
            email_pattern = r"\b[A-Za-z0-9._%+-]+@[A-Za-z0-9.-]+\.[A-Z|a-z]{2,}\b"
            emails = re.findall(email_pattern, response.text)

            # Filter out common non-business emails
            excluded = ["example.com", "test.com", "domain.com", "email.com"]
            for email in emails:
                if not any(excluded_domain in email.lower() for excluded_domain in excluded):
                    return email

        except Exception as e:
            logger.debug(f"Email extraction failed for {website}: {e}")

        return None

    def export_to_excel(
        self, businesses: List[Dict], filename: Optional[str] = None
    ) -> Path:
        """
        Export businesses to Excel file.

        Args:
            businesses: List of business dictionaries
            filename: Output filename (auto-generated if not provided)

        Returns:
            Path to saved Excel file
        """
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"france_businesses_{timestamp}.xlsx"

        filepath = self.output_dir / filename

        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "French Businesses"

        # Define headers
        headers = [
            "Название бизнеса",
            "Категория",
            "Адрес",
            "Телефон",
            "Email",
            "Веб-сайт",
            "Имя владельца",
            "Фамилия владельца",
            "Полное имя владельца",
            "Ссылка Google Maps",
        ]

        # Write headers with styling
        header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )
        header_font = Font(bold=True, color="FFFFFF", size=12)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Write data
        for row_idx, business in enumerate(businesses, 2):
            ws.cell(row=row_idx, column=1, value=business.get("business_name", ""))
            ws.cell(row=row_idx, column=2, value=business.get("category", ""))
            ws.cell(row=row_idx, column=3, value=business.get("address", ""))
            ws.cell(row=row_idx, column=4, value=business.get("phone", ""))
            ws.cell(row=row_idx, column=5, value=business.get("email", ""))
            ws.cell(row=row_idx, column=6, value=business.get("website", ""))
            ws.cell(row=row_idx, column=7, value=business.get("owner_first_name", ""))
            ws.cell(row=row_idx, column=8, value=business.get("owner_last_name", ""))
            ws.cell(row=row_idx, column=9, value=business.get("owner_name", ""))
            ws.cell(row=row_idx, column=10, value=business.get("google_maps_url", ""))

        # Auto-adjust column widths
        for col_idx, header in enumerate(headers, 1):
            max_length = len(header)
            for row_idx in range(2, len(businesses) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[get_column_letter(col_idx)].width = min(
                max_length + 2, 50
            )

        # Save workbook
        wb.save(filepath)
        logger.info(f"Exported {len(businesses)} businesses to {filepath}")

        return filepath


async def main():
    """Main CLI entry point."""
    import argparse

    parser = argparse.ArgumentParser(
        description="Collect French small business data with owner information"
    )
    parser.add_argument(
        "--categories",
        nargs="+",
        default=["petite entreprise", "PME"],
        help="Business categories to search (e.g., 'salon de beauté', 'coiffeur', 'restaurant')",
    )
    parser.add_argument(
        "--location",
        default="France",
        help="Location to search in (default: France)",
    )
    parser.add_argument(
        "--max-results",
        type=int,
        default=20,
        help="Maximum results per category (default: 20)",
    )
    parser.add_argument(
        "--output-dir",
        default="research_results/france",
        help="Output directory for Excel file",
    )
    parser.add_argument(
        "--output-file",
        help="Output Excel filename (auto-generated if not provided)",
    )

    args = parser.parse_args()

    collector = FranceBusinessCollector(output_dir=args.output_dir)

    try:
        # Search for businesses
        logger.info(
            f"Starting search for {len(args.categories)} categories in {args.location}..."
        )
        businesses = await collector.search_businesses(
            categories=args.categories,
            location=args.location,
            max_per_category=args.max_results,
        )

        logger.info(f"Found {len(businesses)} businesses")

        if not businesses:
            logger.warning("No businesses found. Try different categories or location.")
            return

        # Export to Excel
        excel_file = collector.export_to_excel(businesses, args.output_file)
        print(f"\n✅ Success!")
        print(f"   Total businesses: {len(businesses)}")
        print(f"   Excel file: {excel_file}")
        print(f"\n   Statistics:")
        print(f"   - With phone: {sum(1 for b in businesses if b.get('phone'))}")
        print(f"   - With email: {sum(1 for b in businesses if b.get('email'))}")
        print(f"   - With owner name: {sum(1 for b in businesses if b.get('owner_name'))}")

    finally:
        await collector.close()


if __name__ == "__main__":
    asyncio.run(main())
