"""Проверка структуры Excel файла."""
import openpyxl
import sys
from pathlib import Path

excel_file = Path("leads/prague_businesses_20260118_091513.xlsx")
if not excel_file.exists():
    print(f"Файл не найден: {excel_file}")
    sys.exit(1)

wb = openpyxl.load_workbook(excel_file)
ws = wb.active

print(f"Лист: {ws.title}")
print(f"Максимальная строка: {ws.max_row}")
print(f"Максимальная колонка: {ws.max_column}")

if ws.max_row > 0:
    print("\nКолонки (заголовки):")
    headers = []
    for col in ws.iter_cols(min_row=1, max_row=1, max_col=ws.max_column):
        for cell in col:
            if cell.value:
                headers.append(cell.value)
    print(headers)
else:
    print("\nФайл пустой (только заголовки)")
